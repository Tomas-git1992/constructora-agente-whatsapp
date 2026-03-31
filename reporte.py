"""
reporte.py
Genera reportes Excel completos con análisis económico y los envía por WhatsApp via Twilio.
"""
import os
import uuid
import httpx
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Optional

import database as db

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

REPORTES_DIR = Path("/tmp/reportes")
REPORTES_DIR.mkdir(exist_ok=True)

PUBLIC_BASE_URL = os.environ.get(
    "PUBLIC_URL",
    "https://web-production-73e40.up.railway.app"
).rstrip("/")

TWILIO_ACCOUNT_SID   = os.environ.get("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN    = os.environ.get("TWILIO_AUTH_TOKEN", "")
TWILIO_WHATSAPP_FROM = os.environ.get("TWILIO_WHATSAPP_FROM", "")


# ─────────────────────────────────────────────────────────────────────────────
# CONSULTAS A SUPABASE (sin límite, para reportes completos)
# ─────────────────────────────────────────────────────────────────────────────

def _get_movimientos(obra_id=None, fecha_desde=None, fecha_hasta=None):
    c = db.get_client()
    q = (
        c.table("movimientos")
        .select("*, obras(nombre), rubros(nombre), proveedores(nombre)")
        .order("fecha")
    )
    if obra_id:     q = q.eq("obra_id", obra_id)
    if fecha_desde: q = q.gte("fecha", fecha_desde)
    if fecha_hasta: q = q.lte("fecha", fecha_hasta)
    return q.execute().data


def _get_materiales(obra_id=None, fecha_desde=None, fecha_hasta=None):
    c = db.get_client()
    q = (
        c.table("materiales_compra")
        .select("*, obras(nombre), movimientos(fecha, moneda, proveedores(nombre))")
        .order("created_at")
    )
    if obra_id: q = q.eq("obra_id", obra_id)
    rows = q.execute().data
    if fecha_desde or fecha_hasta:
        def en_rango(r):
            f = (r.get("movimientos") or {}).get("fecha", "")
            if fecha_desde and f < fecha_desde: return False
            if fecha_hasta and f > fecha_hasta: return False
            return True
        rows = [r for r in rows if en_rango(r)]
    return rows


def _get_aportes(obra_id=None, fecha_desde=None, fecha_hasta=None):
    c = db.get_client()
    q = (
        c.table("aportes_inversores")
        .select("*, obras(nombre), inversores(nombre)")
        .order("fecha")
    )
    if obra_id:     q = q.eq("obra_id", obra_id)
    if fecha_desde: q = q.gte("fecha", fecha_desde)
    if fecha_hasta: q = q.lte("fecha", fecha_hasta)
    return q.execute().data


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE ESTILO OPENPYXL
# ─────────────────────────────────────────────────────────────────────────────

def _fill(color):
    from openpyxl.styles import PatternFill
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _font(bold=False, color="000000", size=11, italic=False):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size, italic=italic)


def _align(h="left", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _cabecera(ws, row, headers, col_widths=None):
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.fill  = _fill("1F4E79")
        c.font  = _font(bold=True, color="FFFFFF")
        c.alignment = _align("center")
        if col_widths and i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
    ws.row_dimensions[row].height = 20


def _zebra(ws, r, cols):
    if r % 2 == 0:
        for c in range(1, cols + 1):
            ws.cell(r, c).fill = _fill("EBF3FB")


def _seccion(ws, row, txt, cols=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, txt)
    c.font      = _font(bold=True, size=12, color="1F4E79")
    c.fill      = _fill("DEEAF1")
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 1 — RESUMEN EJECUTIVO
# ─────────────────────────────────────────────────────────────────────────────

def _hoja_resumen(ws, movs, mats, aportes, titulo, periodo):
    from openpyxl.utils import get_column_letter

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # Título
    ws.merge_cells("A1:D1")
    ws["A1"].value     = titulo
    ws["A1"].font      = _font(bold=True, size=15, color="1F4E79")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"].value     = f"Período: {periodo}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font      = _font(italic=True, size=10, color="666666")
    ws["A2"].alignment = _align("center")

    row = 4

    # ── Balance por moneda ──────────────────────────────────────────────────
    totales = defaultdict(lambda: {"ing": 0.0, "egr": 0.0})
    for m in movs:
        mon = m.get("moneda", "ARS")
        amt = float(m.get("monto", 0))
        if m.get("tipo") == "ingreso":
            totales[mon]["ing"] += amt
        else:
            totales[mon]["egr"] += amt

    _seccion(ws, row, "💵  BALANCE POR MONEDA"); row += 1
    _cabecera(ws, row, ["Moneda", "Ingresos", "Egresos", "Saldo neto"]); row += 1
    for mon, d in sorted(totales.items()):
        saldo = d["ing"] - d["egr"]
        ws.cell(row, 1, mon)
        ws.cell(row, 2, d["ing"]).number_format = "#,##0.00"
        ws.cell(row, 3, d["egr"]).number_format = "#,##0.00"
        sc = ws.cell(row, 4, saldo)
        sc.number_format = "#,##0.00"
        sc.font = _font(bold=True, color="375623" if saldo >= 0 else "C00000")
        _zebra(ws, row, 4); row += 1
    row += 1

    # ── Saldo de cajas ──────────────────────────────────────────────────────
    _seccion(ws, row, "🏦  SALDO DE CAJAS"); row += 1
    try:
        saldos = db.consultar_saldo_caja()
        if saldos:
            _cabecera(ws, row, ["Obra", "Moneda", "Ingresos", "Egresos", "Saldo"],
                      [28, 12, 16, 16, 16]); row += 1
            for s in saldos:
                ws.cell(row, 1, s.get("obra", ""))
                ws.cell(row, 2, s.get("moneda", ""))
                ws.cell(row, 3, float(s.get("ingresos", 0))).number_format = "#,##0.00"
                ws.cell(row, 4, float(s.get("egresos", 0))).number_format = "#,##0.00"
                ws.cell(row, 5, float(s.get("saldo", 0))).number_format = "#,##0.00"
                _zebra(ws, row, 5); row += 1
        else:
            ws.cell(row, 1, "Sin datos de cajas"); row += 1
    except Exception:
        ws.cell(row, 1, "Vista de cajas no disponible"); row += 1
    row += 1

    # ── Egresos por rubro con % ─────────────────────────────────────────────
    _seccion(ws, row, "📂  EGRESOS POR RUBRO"); row += 1
    _cabecera(ws, row, ["Rubro", "Moneda", "Total egreso", "% del total"]); row += 1
    por_rubro: dict = defaultdict(lambda: defaultdict(float))
    for m in movs:
        if m.get("tipo") == "egreso":
            rub = (m.get("rubros") or {}).get("nombre") or "Sin rubro"
            mon = m.get("moneda", "ARS")
            por_rubro[rub][mon] += float(m.get("monto", 0))
    for rub, monedas in sorted(por_rubro.items()):
        for mon, amt in monedas.items():
            total_e = totales.get(mon, {}).get("egr", 1) or 1
            ws.cell(row, 1, rub)
            ws.cell(row, 2, mon)
            ws.cell(row, 3, amt).number_format = "#,##0.00"
            ws.cell(row, 4, amt / total_e).number_format = "0.0%"
            _zebra(ws, row, 4); row += 1
    row += 1

    # ── Top 10 materiales más costosos ─────────────────────────────────────
    _seccion(ws, row, "🔝  TOP 10 MATERIALES MÁS COSTOSOS"); row += 1
    _cabecera(ws, row, ["Material", "Proveedor", "Total", "Moneda"]); row += 1
    top_mats = sorted(
        [m for m in mats if m.get("precio_total")],
        key=lambda x: float(x.get("precio_total", 0)),
        reverse=True,
    )[:10]
    for m in top_mats:
        mov = m.get("movimientos") or {}
        ws.cell(row, 1, m.get("nombre", ""))
        ws.cell(row, 2, (mov.get("proveedores") or {}).get("nombre", ""))
        ws.cell(row, 3, float(m.get("precio_total", 0))).number_format = "#,##0.00"
        ws.cell(row, 4, mov.get("moneda", ""))
        _zebra(ws, row, 4); row += 1
    if not top_mats:
        ws.cell(row, 1, "Sin materiales con precio registrado"); row += 1
    row += 1

    # ── Aportes de inversores ───────────────────────────────────────────────
    _seccion(ws, row, "🤝  APORTES DE INVERSORES"); row += 1
    _cabecera(ws, row, ["Inversor", "Obra", "Moneda", "Total aportado"]); row += 1
    inv_totales: dict = defaultdict(float)
    inv_info: dict = {}
    for a in aportes:
        inv  = (a.get("inversores") or {}).get("nombre", "Desconocido")
        obra = (a.get("obras") or {}).get("nombre", "")
        mon  = a.get("moneda", "ARS")
        key  = f"{inv}||{obra}||{mon}"
        inv_totales[key] += float(a.get("monto", 0))
        if key not in inv_info:
            inv_info[key] = (inv, obra, mon)
    if inv_totales:
        for key, tot in sorted(inv_totales.items(), key=lambda x: x[1], reverse=True):
            inv_n, obra_n, mon = inv_info[key]
            ws.cell(row, 1, inv_n)
            ws.cell(row, 2, obra_n)
            ws.cell(row, 3, mon)
            ws.cell(row, 4, tot).number_format = "#,##0.00"
            _zebra(ws, row, 4); row += 1
    else:
        ws.cell(row, 1, "Sin aportes registrados"); row += 1


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 2 — MOVIMIENTOS DETALLADOS
# ─────────────────────────────────────────────────────────────────────────────

def _hoja_movimientos(ws, movs):
    headers = ["Fecha", "Obra", "Tipo", "Rubro", "Proveedor",
               "Descripción", "Monto", "Moneda", "Registrado por", "Notas"]
    widths  = [12, 22, 10, 18, 22, 42, 14, 10, 20, 30]
    _cabecera(ws, 1, headers, widths)
    for r, m in enumerate(movs, 2):
        ws.cell(r, 1, m.get("fecha", ""))
        ws.cell(r, 2, (m.get("obras") or {}).get("nombre", ""))
        tipo = m.get("tipo", "")
        ct = ws.cell(r, 3, tipo)
        ct.font = _font(bold=True, color="375623" if tipo == "ingreso" else "C00000")
        ws.cell(r, 4, (m.get("rubros") or {}).get("nombre", ""))
        ws.cell(r, 5, (m.get("proveedores") or {}).get("nombre", ""))
        ws.cell(r, 6, m.get("descripcion", ""))
        ws.cell(r, 7, float(m.get("monto", 0))).number_format = "#,##0.00"
        ws.cell(r, 8, m.get("moneda", ""))
        ws.cell(r, 9, m.get("registrado_por", "") or "")
        ws.cell(r, 10, m.get("notas", "") or "")
        _zebra(ws, r, 10)
    ws.auto_filter.ref = f"A1:J{max(len(movs), 1) + 1}"
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 3 — MATERIALES COMPRADOS
# ─────────────────────────────────────────────────────────────────────────────

def _hoja_materiales(ws, mats):
    headers = ["Fecha", "Obra", "Proveedor", "Material",
               "Cantidad", "Unidad", "Precio Unit.", "Total", "Moneda"]
    widths  = [12, 22, 24, 42, 10, 10, 15, 15, 10]
    _cabecera(ws, 1, headers, widths)
    for r, m in enumerate(mats, 2):
        mov = m.get("movimientos") or {}
        ws.cell(r, 1, mov.get("fecha", ""))
        ws.cell(r, 2, (m.get("obras") or {}).get("nombre", ""))
        ws.cell(r, 3, (mov.get("proveedores") or {}).get("nombre", ""))
        ws.cell(r, 4, m.get("nombre", ""))
        if m.get("cantidad") is not None:
            ws.cell(r, 5, float(m["cantidad"]))
        ws.cell(r, 6, m.get("unidad", "") or "")
        if m.get("precio_unitario") is not None:
            ws.cell(r, 7, float(m["precio_unitario"])).number_format = "#,##0.00"
        if m.get("precio_total") is not None:
            ws.cell(r, 8, float(m["precio_total"])).number_format = "#,##0.00"
        ws.cell(r, 9, mov.get("moneda", ""))
        _zebra(ws, r, 9)
    ws.auto_filter.ref = f"A1:I{max(len(mats), 1) + 1}"
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 4 — APORTES DE INVERSORES
# ─────────────────────────────────────────────────────────────────────────────

def _hoja_inversores(ws, aportes):
    headers = ["Fecha", "Inversor", "Obra", "Tipo", "Monto", "Moneda", "Descripción"]
    widths  = [12, 26, 26, 14, 15, 10, 40]
    _cabecera(ws, 1, headers, widths)
    for r, a in enumerate(aportes, 2):
        ws.cell(r, 1, a.get("fecha", ""))
        ws.cell(r, 2, (a.get("inversores") or {}).get("nombre", ""))
        ws.cell(r, 3, (a.get("obras") or {}).get("nombre", ""))
        ws.cell(r, 4, a.get("tipo", ""))
        ws.cell(r, 5, float(a.get("monto", 0))).number_format = "#,##0.00"
        ws.cell(r, 6, a.get("moneda", ""))
        ws.cell(r, 7, a.get("descripcion", "") or "")
        _zebra(ws, r, 7)
    ws.auto_filter.ref = f"A1:G{max(len(aportes), 1) + 1}"
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# HOJA 5 — EVOLUCIÓN MES A MES
# ─────────────────────────────────────────────────────────────────────────────

def _hoja_mes_a_mes(ws, movs):
    from openpyxl.utils import get_column_letter

    por_mes: dict = defaultdict(lambda: defaultdict(lambda: {"ing": 0.0, "egr": 0.0}))
    for m in movs:
        f = m.get("fecha", "")
        if not f:
            continue
        mes = f[:7]
        mon = m.get("moneda", "ARS")
        amt = float(m.get("monto", 0))
        if m.get("tipo") == "ingreso":
            por_mes[mes][mon]["ing"] += amt
        else:
            por_mes[mes][mon]["egr"] += amt

    monedas = sorted({mon for md in por_mes.values() for mon in md})
    headers = ["Mes"]
    for mon in monedas:
        headers += [f"Ing. {mon}", f"Egr. {mon}", f"Saldo {mon}"]
    widths = [12] + [16] * (len(monedas) * 3)
    _cabecera(ws, 1, headers, widths)

    for r, mes in enumerate(sorted(por_mes.keys()), 2):
        ws.cell(r, 1, mes)
        col = 2
        for mon in monedas:
            d = por_mes[mes].get(mon, {"ing": 0.0, "egr": 0.0})
            saldo = d["ing"] - d["egr"]
            ws.cell(r, col,     d["ing"]).number_format = "#,##0.00"
            ws.cell(r, col + 1, d["egr"]).number_format = "#,##0.00"
            sc = ws.cell(r, col + 2, saldo)
            sc.number_format = "#,##0.00"
            sc.font = _font(bold=True, color="375623" if saldo >= 0 else "C00000")
            col += 3
        _zebra(ws, r, len(headers))
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PRINCIPAL — Generar Excel
# ─────────────────────────────────────────────────────────────────────────────

def generar_reporte_xlsx(
    obra_nombre: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    titulo:      Optional[str] = None,
) -> tuple[str, str]:
    """
    Genera el reporte Excel completo.
    Retorna (token, url_publica).
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl no instalado — agregá 'openpyxl>=3.1.0' a requirements.txt"
        )

    # Resolver obra
    obra_id    = None
    obra_label = "Todas las obras"
    if obra_nombre:
        obj = db.buscar_obra_por_nombre(obra_nombre)
        if obj:
            obra_id    = obj["id"]
            obra_label = obj["nombre"]

    # Obtener datos
    movs    = _get_movimientos(obra_id, fecha_desde, fecha_hasta)
    mats    = _get_materiales(obra_id, fecha_desde, fecha_hasta)
    aportes = _get_aportes(obra_id, fecha_desde, fecha_hasta)

    # Período legible
    if fecha_desde and fecha_hasta:
        periodo = f"{fecha_desde} al {fecha_hasta}"
    elif fecha_desde:
        periodo = f"Desde {fecha_desde}"
    elif fecha_hasta:
        periodo = f"Hasta {fecha_hasta}"
    else:
        periodo = "Todos los períodos"

    titulo_final = titulo or f"Reporte Financiero — {obra_label}"

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    ws1 = wb.create_sheet("📊 Resumen")
    _hoja_resumen(ws1, movs, mats, aportes, titulo_final, periodo)

    ws2 = wb.create_sheet("📋 Movimientos")
    _hoja_movimientos(ws2, movs)

    ws3 = wb.create_sheet("🔩 Materiales")
    _hoja_materiales(ws3, mats)

    ws4 = wb.create_sheet("💰 Inversores")
    _hoja_inversores(ws4, aportes)

    ws5 = wb.create_sheet("📈 Mes a Mes")
    _hoja_mes_a_mes(ws5, movs)

    token = uuid.uuid4().hex[:16]
    path  = REPORTES_DIR / f"{token}.xlsx"
    wb.save(str(path))
    url = f"{PUBLIC_BASE_URL}/reporte/{token}"
    return token, url


# ─────────────────────────────────────────────────────────────────────────────
# ENVÍO POR WHATSAPP (Twilio)
# ─────────────────────────────────────────────────────────────────────────────

def enviar_reporte_whatsapp(token: str, telefono: str) -> None:
    """Envía el Excel como adjunto por WhatsApp via Twilio."""
    url_archivo = f"{PUBLIC_BASE_URL}/reporte/{token}"
    url_api     = (
        f"https://api.twilio.com/2010-04-01/Accounts/"
        f"{TWILIO_ACCOUNT_SID}/Messages.json"
    )
    to = telefono if telefono.startswith("whatsapp:") else f"whatsapp:{telefono}"

    with httpx.Client(timeout=20) as client:
        resp = client.post(
            url_api,
            data={
                "From":     TWILIO_WHATSAPP_FROM,
                "To":       to,
                "Body":     "📊 Acá está tu reporte financiero completo:",
                "MediaUrl": url_archivo,
            },
            auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
        )
    if resp.status_code >= 400:
        raise RuntimeError(f"Error Twilio {resp.status_code}: {resp.text[:300]}")
